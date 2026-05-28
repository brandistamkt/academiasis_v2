"""
app.py - Backend Flask para AcademiaSIS v2
Sistema de gestion de academia - talleres presenciales.
"""
import io
import os
from datetime import date, datetime, timedelta
from decimal import Decimal

import psycopg2
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from database import get_connection, init_db

app = Flask(__name__, static_folder="static", template_folder="templates")


# -------------------- Helpers -------------------- #
def query(sql, params=None, fetch=True, one=False):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(sql, params or ())
    result = None
    if fetch:
        if one:
            row = cur.fetchone()
            result = dict(row) if row else None
        else:
            result = [dict(r) for r in cur.fetchall()]
    conn.commit()
    cur.close()
    conn.close()
    return result


def execute(sql, params=None, returning=False):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(sql, params or ())
    result = None
    if returning:
        row = cur.fetchone()
        result = dict(row) if row else None
    conn.commit()
    cur.close()
    conn.close()
    return result


def to_native(value):
    """Convierte tipos no-serializables a tipos nativos JSON."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def serialize(rows):
    if rows is None:
        return None
    if isinstance(rows, list):
        return [{k: to_native(v) for k, v in row.items()} for row in rows]
    return {k: to_native(v) for k, v in rows.items()}


def parse_date(value, default=None):
    if not value:
        return default
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return default


def get_period_range(period):
    """Devuelve (desde, hasta) segun el periodo: hoy/semana/mes."""
    today = date.today()
    if period == "hoy":
        return today, today
    if period == "semana":
        start = today - timedelta(days=today.weekday())
        return start, today
    if period == "mes":
        start = today.replace(day=1)
        return start, today
    return today, today


# -------------------- Inicializacion BD -------------------- #
@app.before_request
def ensure_db():
    if not getattr(app, "_db_ready", False):
        try:
            init_db()
            app._db_ready = True
        except Exception as e:
            print(f"Error iniciando DB: {e}")


@app.route("/healthz")
def healthz():
    return jsonify({"ok": True, "time": datetime.utcnow().isoformat()})


# -------------------- Frontend -------------------- #
@app.route("/")
def index():
    return render_template("index.html")


# ============================================================
# MODULO: ALUMNOS
# ============================================================
@app.route("/api/alumnos", methods=["GET"])
def listar_alumnos():
    q = request.args.get("q", "").strip()
    if q:
        rows = query(
            """SELECT a.*, COUNT(i.id) AS total_talleres,
                      MAX(f.fecha) AS ultimo_taller
               FROM alumnos a
               LEFT JOIN inscripciones i ON i.alumno_id = a.id
               LEFT JOIN fechas f ON f.id = i.fecha_id
               WHERE LOWER(a.nombre) LIKE LOWER(%s)
                  OR a.dni LIKE %s
                  OR LOWER(COALESCE(a.correo,'')) LIKE LOWER(%s)
               GROUP BY a.id
               ORDER BY a.nombre""",
            (f"%{q}%", f"%{q}%", f"%{q}%"),
        )
    else:
        rows = query(
            """SELECT a.*, COUNT(i.id) AS total_talleres,
                      MAX(f.fecha) AS ultimo_taller
               FROM alumnos a
               LEFT JOIN inscripciones i ON i.alumno_id = a.id
               LEFT JOIN fechas f ON f.id = i.fecha_id
               GROUP BY a.id
               ORDER BY a.creado DESC"""
        )
    return jsonify(serialize(rows))


@app.route("/api/alumnos", methods=["POST"])
def crear_alumno():
    data = request.json or {}
    if not data.get("nombre") or not data.get("dni"):
        return jsonify({"error": "Nombre y DNI son obligatorios"}), 400
    try:
        row = execute(
            """INSERT INTO alumnos (nombre, dni, correo, telefono, edad, profesion, fuente, notas)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
            (
                data["nombre"].strip(),
                data["dni"].strip(),
                data.get("correo"),
                data.get("telefono"),
                data.get("edad") or None,
                data.get("profesion"),
                data.get("fuente") or "Otro",
                data.get("notas"),
            ),
            returning=True,
        )
        return jsonify(serialize(row))
    except psycopg2.errors.UniqueViolation:
        return jsonify({"error": "DNI ya existe"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/alumnos/<int:alumno_id>", methods=["GET"])
def detalle_alumno(alumno_id):
    alumno = query("SELECT * FROM alumnos WHERE id=%s", (alumno_id,), one=True)
    if not alumno:
        return jsonify({"error": "Alumno no encontrado"}), 404
    historial = query(
        """SELECT i.id, i.monto_total, i.monto_pagado, i.estado_pago,
                  i.asistencia, i.notas, i.creado,
                  t.nombre AS taller, f.fecha, f.turno,
                  s.nombre AS sede, f.profesor
           FROM inscripciones i
           JOIN fechas f ON f.id = i.fecha_id
           JOIN talleres t ON t.id = f.taller_id
           LEFT JOIN sedes s ON s.id = f.sede_id
           WHERE i.alumno_id=%s
           ORDER BY f.fecha DESC""",
        (alumno_id,),
    )
    stats = query(
        """SELECT COUNT(*) AS total,
                  COALESCE(SUM(monto_pagado),0) AS pagado,
                  COALESCE(SUM(monto_total - monto_pagado),0) AS saldo,
                  SUM(CASE WHEN asistencia='Asistio' THEN 1 ELSE 0 END) AS asistencias
           FROM inscripciones WHERE alumno_id=%s""",
        (alumno_id,),
        one=True,
    )
    return jsonify({
        "alumno": serialize(alumno),
        "historial": serialize(historial),
        "stats": serialize(stats),
    })


@app.route("/api/alumnos/<int:alumno_id>", methods=["PUT"])
def actualizar_alumno(alumno_id):
    data = request.json or {}
    execute(
        """UPDATE alumnos SET nombre=%s, dni=%s, correo=%s, telefono=%s,
            edad=%s, profesion=%s, fuente=%s, notas=%s WHERE id=%s""",
        (
            data.get("nombre"),
            data.get("dni"),
            data.get("correo"),
            data.get("telefono"),
            data.get("edad") or None,
            data.get("profesion"),
            data.get("fuente"),
            data.get("notas"),
            alumno_id,
        ),
    )
    return jsonify({"ok": True})


@app.route("/api/alumnos/<int:alumno_id>", methods=["DELETE"])
def eliminar_alumno(alumno_id):
    execute("DELETE FROM alumnos WHERE id=%s", (alumno_id,))
    return jsonify({"ok": True})


@app.route("/api/alumnos/buscar_dni")
def buscar_por_dni():
    dni = request.args.get("dni", "").strip()
    if not dni:
        return jsonify(None)
    row = query("SELECT * FROM alumnos WHERE dni=%s", (dni,), one=True)
    return jsonify(serialize(row))


# ============================================================
# MODULO: TALLERES (catalogo)
# ============================================================
@app.route("/api/talleres", methods=["GET"])
def listar_talleres():
    rows = query("SELECT * FROM talleres ORDER BY nombre")
    return jsonify(serialize(rows))


@app.route("/api/talleres", methods=["POST"])
def crear_taller():
    data = request.json or {}
    if not data.get("nombre"):
        return jsonify({"error": "Nombre obligatorio"}), 400
    row = execute(
        """INSERT INTO talleres (nombre, duracion_dias, precio_base, descripcion)
           VALUES (%s,%s,%s,%s) RETURNING *""",
        (data["nombre"], data.get("duracion_dias", 1),
         data.get("precio_base", 0), data.get("descripcion")),
        returning=True,
    )
    return jsonify(serialize(row))


@app.route("/api/talleres/<int:tid>", methods=["DELETE"])
def eliminar_taller(tid):
    execute("DELETE FROM talleres WHERE id=%s", (tid,))
    return jsonify({"ok": True})


# ============================================================
# MODULO: SEDES
# ============================================================
@app.route("/api/sedes", methods=["GET"])
def listar_sedes():
    rows = query("SELECT * FROM sedes ORDER BY nombre")
    return jsonify(serialize(rows))


@app.route("/api/sedes", methods=["POST"])
def crear_sede():
    data = request.json or {}
    if not data.get("nombre"):
        return jsonify({"error": "Nombre obligatorio"}), 400
    row = execute(
        "INSERT INTO sedes (nombre, direccion) VALUES (%s,%s) RETURNING *",
        (data["nombre"], data.get("direccion")),
        returning=True,
    )
    return jsonify(serialize(row))


@app.route("/api/sedes/<int:sid>", methods=["DELETE"])
def eliminar_sede(sid):
    execute("DELETE FROM sedes WHERE id=%s", (sid,))
    return jsonify({"ok": True})


# ============================================================
# MODULO: FECHAS PROGRAMADAS
# ============================================================
@app.route("/api/fechas", methods=["GET"])
def listar_fechas():
    taller_id = request.args.get("taller_id")
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    where, params = [], []
    if taller_id:
        where.append("f.taller_id=%s")
        params.append(taller_id)
    if desde:
        where.append("f.fecha >= %s")
        params.append(desde)
    if hasta:
        where.append("f.fecha <= %s")
        params.append(hasta)
    sql = """
        SELECT f.*, t.nombre AS taller, s.nombre AS sede,
               COUNT(i.id) AS inscritos,
               COALESCE(SUM(i.monto_pagado),0) AS ingresos
        FROM fechas f
        JOIN talleres t ON t.id = f.taller_id
        LEFT JOIN sedes s ON s.id = f.sede_id
        LEFT JOIN inscripciones i ON i.fecha_id = f.id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY f.id, t.nombre, s.nombre ORDER BY f.fecha DESC"
    rows = query(sql, params)
    return jsonify(serialize(rows))


@app.route("/api/fechas", methods=["POST"])
def crear_fecha():
    data = request.json or {}
    if not data.get("taller_id") or not data.get("fecha") or not data.get("turno"):
        return jsonify({"error": "Taller, fecha y turno son obligatorios"}), 400
    row = execute(
        """INSERT INTO fechas (taller_id, fecha, turno, sede_id, profesor, cupo)
           VALUES (%s,%s,%s,%s,%s,%s) RETURNING *""",
        (
            data["taller_id"],
            data["fecha"],
            data["turno"],
            data.get("sede_id") or None,
            data.get("profesor"),
            data.get("cupo", 20),
        ),
        returning=True,
    )
    return jsonify(serialize(row))


@app.route("/api/fechas/<int:fid>", methods=["DELETE"])
def eliminar_fecha(fid):
    execute("DELETE FROM fechas WHERE id=%s", (fid,))
    return jsonify({"ok": True})


@app.route("/api/fechas/<int:fid>/detalle")
def detalle_fecha(fid):
    fecha = query(
        """SELECT f.*, t.nombre AS taller, s.nombre AS sede
           FROM fechas f
           JOIN talleres t ON t.id=f.taller_id
           LEFT JOIN sedes s ON s.id=f.sede_id
           WHERE f.id=%s""",
        (fid,), one=True,
    )
    if not fecha:
        return jsonify({"error": "Fecha no encontrada"}), 404
    inscritos = query(
        """SELECT i.id, i.asistencia, i.monto_total, i.monto_pagado,
                  i.estado_pago, a.nombre, a.dni
           FROM inscripciones i
           JOIN alumnos a ON a.id = i.alumno_id
           WHERE i.fecha_id=%s
           ORDER BY a.nombre""",
        (fid,),
    )
    insumos = query("SELECT * FROM insumos ORDER BY nombre")
    return jsonify({
        "fecha": serialize(fecha),
        "inscritos": serialize(inscritos),
        "insumos": serialize(insumos),
    })


# ============================================================
# MODULO: INSCRIPCIONES
# ============================================================
@app.route("/api/inscripciones", methods=["GET"])
def listar_inscripciones():
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    taller_id = request.args.get("taller_id")
    estado = request.args.get("estado")
    where, params = [], []
    if desde:
        where.append("f.fecha >= %s")
        params.append(desde)
    if hasta:
        where.append("f.fecha <= %s")
        params.append(hasta)
    if taller_id:
        where.append("f.taller_id=%s")
        params.append(taller_id)
    if estado:
        where.append("i.estado_pago=%s")
        params.append(estado)
    sql = """
        SELECT i.*, a.nombre AS alumno, a.dni, t.nombre AS taller,
               f.fecha, f.turno, s.nombre AS sede
        FROM inscripciones i
        JOIN alumnos a ON a.id=i.alumno_id
        JOIN fechas f ON f.id=i.fecha_id
        JOIN talleres t ON t.id=f.taller_id
        LEFT JOIN sedes s ON s.id=f.sede_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY i.creado DESC"
    rows = query(sql, params)
    return jsonify(serialize(rows))


@app.route("/api/inscripciones", methods=["POST"])
def crear_inscripcion():
    data = request.json or {}
    if not data.get("alumno_id") or not data.get("fecha_id"):
        return jsonify({"error": "Alumno y fecha son obligatorios"}), 400
    monto_total = float(data.get("monto_total") or 0)
    monto_pagado = float(data.get("monto_pagado") or 0)
    if monto_pagado >= monto_total and monto_total > 0:
        estado = "Pagado"
    elif monto_pagado > 0:
        estado = "Adelanto"
    else:
        estado = "Pendiente"
    row = execute(
        """INSERT INTO inscripciones
            (alumno_id, fecha_id, monto_total, monto_pagado, metodo_pago,
             estado_pago, promocion, tipo_descuento, notas)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
        (
            data["alumno_id"], data["fecha_id"], monto_total, monto_pagado,
            data.get("metodo_pago"), estado,
            data.get("promocion"), data.get("tipo_descuento"),
            data.get("notas"),
        ),
        returning=True,
    )
    return jsonify(serialize(row))


@app.route("/api/inscripciones/<int:iid>", methods=["PUT"])
def actualizar_inscripcion(iid):
    data = request.json or {}
    monto_total = float(data.get("monto_total") or 0)
    monto_pagado = float(data.get("monto_pagado") or 0)
    if monto_pagado >= monto_total and monto_total > 0:
        estado = "Pagado"
    elif monto_pagado > 0:
        estado = "Adelanto"
    else:
        estado = "Pendiente"
    execute(
        """UPDATE inscripciones SET monto_total=%s, monto_pagado=%s,
           metodo_pago=%s, estado_pago=%s, promocion=%s, tipo_descuento=%s,
           asistencia=COALESCE(%s, asistencia), notas=%s WHERE id=%s""",
        (
            monto_total, monto_pagado, data.get("metodo_pago"), estado,
            data.get("promocion"), data.get("tipo_descuento"),
            data.get("asistencia"), data.get("notas"), iid,
        ),
    )
    return jsonify({"ok": True})


@app.route("/api/inscripciones/<int:iid>", methods=["DELETE"])
def eliminar_inscripcion(iid):
    execute("DELETE FROM inscripciones WHERE id=%s", (iid,))
    return jsonify({"ok": True})


# ============================================================
# MODULO: CIERRE DE CLASE
# ============================================================
@app.route("/api/fechas/<int:fid>/cerrar", methods=["POST"])
def cerrar_clase(fid):
    data = request.json or {}
    # Actualizar metadata de la fecha
    execute(
        """UPDATE fechas SET estado='cerrada',
            valoracion=%s, que_paso=%s, que_mejorar=%s, que_funciono=%s,
            cerrado=CURRENT_TIMESTAMP WHERE id=%s""",
        (
            data.get("valoracion"),
            ",".join(data.get("que_paso", [])) if isinstance(data.get("que_paso"), list) else data.get("que_paso"),
            data.get("que_mejorar"),
            data.get("que_funciono"),
            fid,
        ),
    )
    # Asistencias
    for a in data.get("asistencias", []):
        execute(
            "UPDATE inscripciones SET asistencia=%s WHERE id=%s",
            (a.get("asistencia"), a.get("id")),
        )
    # Consumos de insumos -> descontar stock
    for c in data.get("consumos", []):
        if not c.get("insumo_id") or float(c.get("cantidad") or 0) <= 0:
            continue
        execute(
            "INSERT INTO consumos (fecha_id, insumo_id, cantidad) VALUES (%s,%s,%s)",
            (fid, c["insumo_id"], c["cantidad"]),
        )
        execute(
            "UPDATE insumos SET stock_actual = GREATEST(stock_actual - %s, 0) WHERE id=%s",
            (c["cantidad"], c["insumo_id"]),
        )
    return jsonify({"ok": True})


# ============================================================
# MODULO: GASTOS / FINANZAS
# ============================================================
@app.route("/api/gastos", methods=["GET"])
def listar_gastos():
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    limit = request.args.get("limit")
    where, params = [], []
    if desde:
        where.append("fecha >= %s")
        params.append(desde)
    if hasta:
        where.append("fecha <= %s")
        params.append(hasta)
    sql = "SELECT * FROM gastos"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY fecha DESC, id DESC"
    if limit:
        sql += f" LIMIT {int(limit)}"
    rows = query(sql, params)
    return jsonify(serialize(rows))


@app.route("/api/gastos", methods=["POST"])
def crear_gasto():
    data = request.json or {}
    if not data.get("descripcion") or not data.get("monto") or not data.get("fecha"):
        return jsonify({"error": "Descripcion, monto y fecha son obligatorios"}), 400
    row = execute(
        """INSERT INTO gastos (descripcion, categoria, monto, fecha,
            metodo_pago, proveedor, comprobante, notas)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
        (
            data["descripcion"], data.get("categoria"), data["monto"],
            data["fecha"], data.get("metodo_pago"), data.get("proveedor"),
            data.get("comprobante"), data.get("notas"),
        ),
        returning=True,
    )
    return jsonify(serialize(row))


@app.route("/api/gastos/<int:gid>", methods=["PUT"])
def actualizar_gasto(gid):
    data = request.json or {}
    execute(
        """UPDATE gastos SET descripcion=%s, categoria=%s, monto=%s,
            fecha=%s, metodo_pago=%s, proveedor=%s, comprobante=%s, notas=%s
           WHERE id=%s""",
        (
            data.get("descripcion"), data.get("categoria"), data.get("monto"),
            data.get("fecha"), data.get("metodo_pago"), data.get("proveedor"),
            data.get("comprobante"), data.get("notas"), gid,
        ),
    )
    return jsonify({"ok": True})


@app.route("/api/gastos/<int:gid>", methods=["DELETE"])
def eliminar_gasto(gid):
    execute("DELETE FROM gastos WHERE id=%s", (gid,))
    return jsonify({"ok": True})


@app.route("/api/finanzas/resumen")
def finanzas_resumen():
    desde = request.args.get("desde") or (date.today() - timedelta(days=30)).isoformat()
    hasta = request.args.get("hasta") or date.today().isoformat()
    ingresos = query(
        """SELECT COALESCE(SUM(i.monto_pagado),0) AS total
           FROM inscripciones i
           JOIN fechas f ON f.id=i.fecha_id
           WHERE f.fecha BETWEEN %s AND %s""",
        (desde, hasta), one=True,
    )
    gastos = query(
        "SELECT COALESCE(SUM(monto),0) AS total FROM gastos WHERE fecha BETWEEN %s AND %s",
        (desde, hasta), one=True,
    )
    por_categoria = query(
        """SELECT categoria, COALESCE(SUM(monto),0) AS total
           FROM gastos WHERE fecha BETWEEN %s AND %s
           GROUP BY categoria ORDER BY total DESC""",
        (desde, hasta),
    )
    descuentos = query(
        """SELECT tipo_descuento AS promo, COUNT(*) AS usos,
                  COALESCE(SUM(monto_total - monto_pagado),0) AS estimado
           FROM inscripciones
           WHERE tipo_descuento IS NOT NULL AND tipo_descuento <> ''
           GROUP BY tipo_descuento""",
    )
    total_ing = float(ingresos["total"] or 0)
    total_gas = float(gastos["total"] or 0)
    return jsonify({
        "ingresos": total_ing,
        "gastos": total_gas,
        "utilidad": total_ing - total_gas,
        "por_categoria": serialize(por_categoria),
        "descuentos": serialize(descuentos),
    })


# ============================================================
# MODULO: RENTABILIDAD
# ============================================================
@app.route("/api/rentabilidad")
def rentabilidad():
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    where, params = [], []
    if desde:
        where.append("f.fecha >= %s")
        params.append(desde)
    if hasta:
        where.append("f.fecha <= %s")
        params.append(hasta)
    sql = """
        SELECT t.id, t.nombre AS taller,
               COUNT(i.id) AS inscritos,
               COALESCE(SUM(i.monto_pagado),0) AS ingresos,
               COALESCE(SUM(i.monto_total - i.monto_pagado),0) AS por_cobrar,
               SUM(CASE WHEN i.asistencia='Asistio' THEN 1 ELSE 0 END) AS asistencias,
               CASE WHEN COUNT(i.id) > 0
                    THEN COALESCE(SUM(i.monto_pagado),0)/COUNT(i.id)
                    ELSE 0 END AS ticket_promedio
        FROM talleres t
        LEFT JOIN fechas f ON f.taller_id=t.id
        LEFT JOIN inscripciones i ON i.fecha_id=f.id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY t.id, t.nombre ORDER BY ingresos DESC"
    rows = query(sql, params)
    return jsonify(serialize(rows))


# ============================================================
# MODULO: INVENTARIO
# ============================================================
@app.route("/api/insumos", methods=["GET"])
def listar_insumos():
    rows = query("SELECT * FROM insumos ORDER BY nombre")
    return jsonify(serialize(rows))


@app.route("/api/insumos", methods=["POST"])
def crear_insumo():
    data = request.json or {}
    if not data.get("nombre"):
        return jsonify({"error": "Nombre obligatorio"}), 400
    row = execute(
        """INSERT INTO insumos (nombre, unidad, stock_actual, stock_minimo,
            consumo_clase, precio_unitario)
           VALUES (%s,%s,%s,%s,%s,%s) RETURNING *""",
        (
            data["nombre"], data.get("unidad", "uds"),
            data.get("stock_actual", 0), data.get("stock_minimo", 0),
            data.get("consumo_clase", 0), data.get("precio_unitario", 0),
        ),
        returning=True,
    )
    return jsonify(serialize(row))


@app.route("/api/insumos/<int:iid>", methods=["PUT"])
def actualizar_insumo(iid):
    data = request.json or {}
    execute(
        """UPDATE insumos SET nombre=%s, unidad=%s, stock_actual=%s,
            stock_minimo=%s, consumo_clase=%s, precio_unitario=%s WHERE id=%s""",
        (
            data.get("nombre"), data.get("unidad"),
            data.get("stock_actual"), data.get("stock_minimo"),
            data.get("consumo_clase"), data.get("precio_unitario"), iid,
        ),
    )
    return jsonify({"ok": True})


@app.route("/api/insumos/<int:iid>/agregar", methods=["POST"])
def agregar_stock(iid):
    data = request.json or {}
    cantidad = float(data.get("cantidad") or 0)
    execute(
        "UPDATE insumos SET stock_actual = stock_actual + %s WHERE id=%s",
        (cantidad, iid),
    )
    return jsonify({"ok": True})


@app.route("/api/insumos/<int:iid>", methods=["DELETE"])
def eliminar_insumo(iid):
    execute("DELETE FROM insumos WHERE id=%s", (iid,))
    return jsonify({"ok": True})


# ============================================================
# MODULO: DASHBOARD
# ============================================================
@app.route("/api/dashboard")
def dashboard():
    period = request.args.get("period", "hoy")
    desde, hasta = get_period_range(period)
    today = date.today()

    # Inscritos en el periodo
    insc_periodo = query(
        """SELECT COUNT(*) AS c FROM inscripciones i
           JOIN fechas f ON f.id=i.fecha_id
           WHERE i.creado::date BETWEEN %s AND %s""",
        (desde, hasta), one=True,
    )
    total_alumnos = query("SELECT COUNT(*) AS c FROM alumnos", one=True)

    ingresos = query(
        """SELECT COALESCE(SUM(i.monto_pagado),0) AS total
           FROM inscripciones i WHERE i.creado::date BETWEEN %s AND %s""",
        (desde, hasta), one=True,
    )
    gastos = query(
        "SELECT COALESCE(SUM(monto),0) AS total FROM gastos WHERE fecha BETWEEN %s AND %s",
        (desde, hasta), one=True,
    )
    ing_v = float(ingresos["total"] or 0)
    gas_v = float(gastos["total"] or 0)
    utilidad = ing_v - gas_v
    margen = (utilidad / ing_v * 100) if ing_v > 0 else 0

    # Alertas
    stock_critico = query(
        "SELECT * FROM insumos WHERE stock_actual <= stock_minimo OR (consumo_clase>0 AND stock_actual/consumo_clase < 3)",
    )
    pagos_pend = query(
        "SELECT COUNT(*) AS c FROM inscripciones WHERE estado_pago <> 'Pagado'",
        one=True,
    )
    talleres_mañana = query(
        """SELECT COUNT(*) AS c FROM fechas WHERE fecha = %s AND estado='activa'""",
        (today + timedelta(days=1),), one=True,
    )

    # Grafica 7 dias
    chart_data = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        ing_dia = query(
            """SELECT COALESCE(SUM(monto_pagado),0) AS t FROM inscripciones
               WHERE creado::date = %s""",
            (d,), one=True,
        )
        gas_dia = query(
            "SELECT COALESCE(SUM(monto),0) AS t FROM gastos WHERE fecha=%s",
            (d,), one=True,
        )
        chart_data.append({
            "fecha": d.isoformat(),
            "ingresos": float(ing_dia["t"] or 0),
            "gastos": float(gas_dia["t"] or 0),
        })

    # Top 5 cursos
    top_cursos = query(
        """SELECT t.nombre AS taller, COUNT(i.id) AS inscritos,
                  COALESCE(SUM(i.monto_pagado),0) AS ingresos,
                  SUM(CASE WHEN i.asistencia='Asistio' THEN 1 ELSE 0 END) AS asistencias,
                  CASE WHEN COUNT(i.id)>0
                       THEN COALESCE(SUM(i.monto_pagado),0)/COUNT(i.id)
                       ELSE 0 END AS ticket
           FROM talleres t
           LEFT JOIN fechas f ON f.taller_id=t.id
           LEFT JOIN inscripciones i ON i.fecha_id=f.id
           GROUP BY t.id, t.nombre
           ORDER BY ingresos DESC LIMIT 5""",
    )

    # Fuente captacion
    fuentes_rows = query(
        "SELECT fuente, COUNT(*) AS c FROM alumnos GROUP BY fuente",
    )
    total_f = sum(int(f["c"]) for f in fuentes_rows) or 1
    fuentes = [{"fuente": f["fuente"], "cantidad": int(f["c"]),
                "porcentaje": round(int(f["c"]) * 100 / total_f, 1)}
               for f in fuentes_rows]

    # Tasa retorno (alumnos con >1 inscripcion)
    retorno = query(
        """SELECT
             COUNT(*) FILTER (WHERE total>1) AS recurrentes,
             COUNT(*) AS total
           FROM (SELECT alumno_id, COUNT(*) AS total FROM inscripciones GROUP BY alumno_id) x""",
        one=True,
    )
    rec = int(retorno["recurrentes"] or 0)
    tot = int(retorno["total"] or 0)
    tasa_retorno = round(rec * 100 / tot, 1) if tot > 0 else 0

    # Proximos talleres
    proximos = query(
        """SELECT f.id, f.fecha, f.turno, f.profesor, f.cupo, f.estado,
                  t.nombre AS taller, s.nombre AS sede,
                  COUNT(i.id) AS inscritos
           FROM fechas f
           JOIN talleres t ON t.id=f.taller_id
           LEFT JOIN sedes s ON s.id=f.sede_id
           LEFT JOIN inscripciones i ON i.fecha_id=f.id
           WHERE f.fecha >= %s AND f.estado='activa'
           GROUP BY f.id, t.nombre, s.nombre
           ORDER BY f.fecha ASC LIMIT 10""",
        (today,),
    )

    # Inscripciones por taller (resumen)
    insc_por_taller = query(
        """SELECT f.id, t.nombre AS taller, f.fecha, f.turno, f.cupo,
                  COUNT(i.id) AS inscritos,
                  SUM(CASE WHEN i.asistencia='Asistio' THEN 1 ELSE 0 END) AS asistencias
           FROM fechas f
           JOIN talleres t ON t.id=f.taller_id
           LEFT JOIN inscripciones i ON i.fecha_id=f.id
           GROUP BY f.id, t.nombre
           ORDER BY f.fecha DESC LIMIT 10""",
    )

    # Reprogramados
    reprogramados = query(
        """SELECT a.nombre, a.dni, t.nombre AS taller, f.fecha
           FROM inscripciones i
           JOIN alumnos a ON a.id=i.alumno_id
           JOIN fechas f ON f.id=i.fecha_id
           JOIN talleres t ON t.id=f.taller_id
           WHERE i.asistencia='Reprogramo'""",
    )

    # Seguimientos pendientes
    seguimientos = query(
        """SELECT i.id, a.nombre, a.dni, t.nombre AS taller, f.fecha,
                  i.estado_pago, i.asistencia,
                  i.monto_total, i.monto_pagado
           FROM inscripciones i
           JOIN alumnos a ON a.id=i.alumno_id
           JOIN fechas f ON f.id=i.fecha_id
           JOIN talleres t ON t.id=f.taller_id
           WHERE i.estado_pago <> 'Pagado'
              OR i.asistencia = 'Pendiente'
           ORDER BY f.fecha DESC LIMIT 20""",
    )

    return jsonify({
        "metrics": {
            "inscritos_periodo": int(insc_periodo["c"]),
            "total_alumnos": int(total_alumnos["c"]),
            "ingresos": ing_v,
            "gastos": gas_v,
            "utilidad": utilidad,
            "margen": round(margen, 1),
        },
        "alertas": {
            "stock_critico": len(stock_critico),
            "pagos_pendientes": int(pagos_pend["c"]),
            "talleres_mañana": int(talleres_mañana["c"]),
        },
        "chart": chart_data,
        "top_cursos": serialize(top_cursos),
        "fuentes": fuentes,
        "tasa_retorno": tasa_retorno,
        "stock_critico": serialize(stock_critico),
        "proximos": serialize(proximos),
        "insc_por_taller": serialize(insc_por_taller),
        "reprogramados": serialize(reprogramados),
        "seguimientos": serialize(seguimientos),
    })


# ============================================================
# MODULO: REPORTES - EXPORTACION EXCEL
# ============================================================
@app.route("/api/reportes/excel")
def exportar_excel():
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    where, params = [], []
    if desde:
        where.append("f.fecha >= %s")
        params.append(desde)
    if hasta:
        where.append("f.fecha <= %s")
        params.append(hasta)
    sql_insc = """
        SELECT a.nombre, a.dni, a.correo, a.telefono, a.edad, a.profesion, a.fuente,
               t.nombre AS taller, f.fecha, f.turno, f.profesor, s.nombre AS sede,
               i.estado_pago, i.monto_total, i.monto_pagado,
               (i.monto_total - i.monto_pagado) AS saldo,
               i.metodo_pago, i.promocion, i.tipo_descuento, i.asistencia, i.notas
        FROM inscripciones i
        JOIN alumnos a ON a.id=i.alumno_id
        JOIN fechas f ON f.id=i.fecha_id
        JOIN talleres t ON t.id=f.taller_id
        LEFT JOIN sedes s ON s.id=f.sede_id
    """
    if where:
        sql_insc += " WHERE " + " AND ".join(where)
    sql_insc += " ORDER BY f.fecha DESC"
    inscripciones = query(sql_insc, params)

    # Gastos
    sql_g = "SELECT descripcion, categoria, monto, fecha, metodo_pago, comprobante, proveedor, notas FROM gastos"
    g_where, g_params = [], []
    if desde:
        g_where.append("fecha >= %s")
        g_params.append(desde)
    if hasta:
        g_where.append("fecha <= %s")
        g_params.append(hasta)
    if g_where:
        sql_g += " WHERE " + " AND ".join(g_where)
    sql_g += " ORDER BY fecha DESC"
    gastos = query(sql_g, g_params)

    # Construir Excel
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Inscripciones"
    headers1 = [
        "Nombre", "DNI", "Correo", "Telefono", "Edad", "Profesion", "Fuente",
        "Taller", "Fecha", "Turno", "Profesor", "Sede",
        "Estado Pago", "Monto Total", "Monto Pagado", "Saldo",
        "Metodo Pago", "Promocion", "Tipo Descuento", "Asistencia", "Anotaciones",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D3748")
        cell.alignment = Alignment(horizontal="center")
    for r in inscripciones:
        ws1.append([
            r.get("nombre"), r.get("dni"), r.get("correo"), r.get("telefono"),
            r.get("edad"), r.get("profesion"), r.get("fuente"),
            r.get("taller"),
            r.get("fecha").isoformat() if r.get("fecha") else "",
            r.get("turno"), r.get("profesor"), r.get("sede"),
            r.get("estado_pago"),
            float(r.get("monto_total") or 0),
            float(r.get("monto_pagado") or 0),
            float(r.get("saldo") or 0),
            r.get("metodo_pago"), r.get("promocion"), r.get("tipo_descuento"),
            r.get("asistencia"), r.get("notas"),
        ])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    ws2 = wb.create_sheet("Gastos")
    headers2 = ["Descripcion", "Categoria", "Monto", "Fecha", "Metodo Pago",
                "Comprobante", "Proveedor", "Notas"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D3748")
        cell.alignment = Alignment(horizontal="center")
    for g in gastos:
        ws2.append([
            g.get("descripcion"), g.get("categoria"),
            float(g.get("monto") or 0),
            g.get("fecha").isoformat() if g.get("fecha") else "",
            g.get("metodo_pago"), g.get("comprobante"),
            g.get("proveedor"), g.get("notas"),
        ])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"reporte_academia_{date.today().isoformat()}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
